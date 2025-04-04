import subprocess
import sys

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

import tkinter as tk
from tkinter import filedialog
from collections import deque

def get_neighbors(r, c, max_row, max_col):
    neighbors = []
    if r > 1: neighbors.append((r - 1, c))
    if r < max_row: neighbors.append((r + 1, c))
    if c > 1: neighbors.append((r, c - 1))
    if c < max_col: neighbors.append((r, c + 1))
    return neighbors

def flood_fill(sheet, start_row, start_col, max_row, max_col, visited):
    queue = deque()
    queue.append((start_row, start_col))
    visited.add((start_row, start_col))
    
    min_r, max_r = start_row, start_row
    min_c, max_c = start_col, start_col
    
    while queue:
        r, c = queue.popleft()
        min_r = min(min_r, r)
        max_r = max(max_r, r)
        min_c = min(min_c, c)
        max_c = max(max_c, c)
        
        for nr, nc in get_neighbors(r, c, max_row, max_col):
            if (nr, nc) not in visited:
                if sheet.cell(nr, nc).value is not None:
                    visited.add((nr, nc))
                    queue.append((nr, nc))
    return min_r, max_r, min_c, max_c

def merge_boxes(boxes, row_tol=1, col_tol=1):
    merged = boxes[:]
    merged_changed = True

    while merged_changed:
        merged_changed = False
        new_merged = []
        skip = [False] * len(merged)
        for i in range(len(merged)):
            if skip[i]:
                continue
            box_a = merged[i]
            a_min_r, a_max_r, a_min_c, a_max_c = box_a
            current_box = box_a
            for j in range(i+1, len(merged)):
                if skip[j]:
                    continue
                box_b = merged[j]
                b_min_r, b_max_r, b_min_c, b_max_c = box_b
                vertical_gap = max(0, max(b_min_r - a_max_r - 1, a_min_r - b_max_r - 1))
                horizontal_gap = max(0, max(b_min_c - a_max_c - 1, a_min_c - b_max_c - 1))
                
                if vertical_gap <= row_tol and not (b_max_c < a_min_c - col_tol or b_min_c > a_max_c + col_tol):
                    current_box = (
                        min(current_box[0], b_min_r),
                        max(current_box[1], b_max_r),
                        min(current_box[2], b_min_c),
                        max(current_box[3], b_max_c)
                    )
                    skip[j] = True
                    merged_changed = True
                elif horizontal_gap <= col_tol and not (b_max_r < a_min_r - row_tol or b_min_r > a_max_r + row_tol):
                    current_box = (
                        min(current_box[0], b_min_r),
                        max(current_box[1], b_max_r),
                        min(current_box[2], b_min_c),
                        max(current_box[3], b_max_c)
                    )
                    skip[j] = True
                    merged_changed = True
            new_merged.append(current_box)
        merged = new_merged
    merged.sort(key=lambda b: (b[0], b[2]))
    return merged

def process_merged_cells(sheet):
    merged_info = []
    for merged_range in sheet.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        top_left_value = sheet.cell(min_row, min_col).value
        merged_info.append((min_row, max_row, min_col, max_col, top_left_value))
    return merged_info

def separate_tables_in_excel(file_path, source_sheet="Sheet1", output_file="output.xlsx",
                            row_tol=1, col_tol=1):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[source_sheet]
    
    merged_info = process_merged_cells(sheet)
    
    visited = set()
    boxes = []
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            if (r, c) not in visited and sheet.cell(r, c).value is not None:
                box = flood_fill(sheet, r, c, sheet.max_row, sheet.max_column, visited)
                boxes.append(box)
    
    merged_boxes = merge_boxes(boxes, row_tol=row_tol, col_tol=col_tol)

    for idx, (min_r, max_r, min_c, max_c) in enumerate(merged_boxes, start=1):
        new_sheet = wb.create_sheet(title=f"Table_{idx}")
        for new_row, orig_row in enumerate(range(min_r, max_r + 1), start=1):
            for new_col, orig_col in enumerate(range(min_c, max_c + 1), start=1):
                cell_value = sheet.cell(orig_row, orig_col).value
                
                for m_min_r, m_max_r, m_min_c, m_max_c, m_value in merged_info:
                    if m_min_r <= orig_row <= m_max_r and m_min_c <= orig_col <= m_max_c:
                        cell_value = m_value
                        break
                
                new_sheet.cell(new_row, new_col, cell_value)
    wb.remove(sheet)
    wb.save(output_file)
    print(f"Separated {len(merged_boxes)} tables into individual sheets in '{output_file}'")

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
    )
    
    if file_path:
        save_path = filedialog.asksaveasfilename(
            title="Save Output File",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        
        if save_path:
            separate_tables_in_excel(file_path, output_file=save_path)
        else:
            print("No save location selected")
    else:
        print("No file selected")