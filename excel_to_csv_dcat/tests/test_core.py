"""Unit tests for the core functionality."""
import os
import unittest
from io import BytesIO
import pandas as pd
import openpyxl
from rdflib import Graph
from excel_to_csv_dcat.core import (
    extract_tables_from_excel,
    process_excel_in_memory,
    generate_dcat_metadata,
    merge_boxes,
    flood_fill
)

class TestTableDetection(unittest.TestCase):
    def setUp(self):
        """Set up test fixtures."""
        # Create a simple Excel file in memory for testing
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        
        # Create a simple table
        self.ws['A1'] = 'Header 1'
        self.ws['B1'] = 'Header 2'
        self.ws['A2'] = 'Data 1'
        self.ws['B2'] = 'Data 2'
        
        # Create a second table with a gap
        self.ws['D1'] = 'Table 2'
        self.ws['D2'] = 'Value 1'
        self.ws['E2'] = 'Value 2'
        
        # Save to BytesIO
        self.excel_buffer = BytesIO()
        self.wb.save(self.excel_buffer)
        self.excel_bytes = self.excel_buffer.getvalue()
    
    def test_extract_tables(self):
        """Test table extraction from Excel."""
        tables = extract_tables_from_excel(self.excel_bytes)
        self.assertEqual(len(tables), 2, "Should detect two tables")
        
        # Check first table
        name1, buf1 = tables[0]
        df1 = pd.read_csv(buf1)
        self.assertEqual(df1.shape, (2, 2), "First table should be 2x2")
        
        # Check second table
        name2, buf2 = tables[1]
        df2 = pd.read_csv(buf2)
        self.assertEqual(df2.shape, (2, 2), "Second table should be 2x2")

    def test_merge_boxes(self):
        """Test merging of nearby table boxes."""
        boxes = [
            (1, 2, 1, 2),  # First box
            (4, 5, 1, 2),  # Second box, vertically separated
            (1, 2, 4, 5),  # Third box, horizontally separated
        ]
        
        # Test with default tolerance
        merged = merge_boxes(boxes)
        self.assertEqual(len(merged), 2, "Should merge vertically adjacent boxes")
        
        # Test with zero tolerance
        merged_no_tol = merge_boxes(boxes, row_tol=0, col_tol=0)
        self.assertEqual(len(merged_no_tol), 3, "Should not merge boxes with zero tolerance")

    def test_merge_boxes_no_overlap(self):
        """Test merge_boxes with non-overlapping boxes."""
        boxes = [(1, 2, 1, 2), (3, 4, 3, 4)]
        merged = merge_boxes(boxes)
        self.assertEqual(len(merged), 2, "Should not merge non-overlapping boxes.")

    def test_merge_boxes_with_overlap(self):
        """Test merge_boxes with overlapping boxes."""
        boxes = [(1, 2, 1, 2), (2, 3, 2, 3)]
        merged = merge_boxes(boxes)
        self.assertEqual(len(merged), 1, "Should merge overlapping boxes.")
        self.assertEqual(merged[0], (1, 3, 1, 3), "Merged box should have correct boundaries.")

    def test_flood_fill(self):
        """Test flood fill table detection."""
        visited = set()
        min_r, max_r, min_c, max_c = flood_fill(
            self.ws, 1, 1, self.ws.max_row, self.ws.max_column, visited
        )
        
        self.assertEqual(min_r, 1)
        self.assertEqual(max_r, 2)
        self.assertEqual(min_c, 1)
        self.assertEqual(max_c, 2)
        self.assertEqual(len(visited), 4, "Should visit all cells in first table")

    def test_flood_fill_empty_sheet(self):
        """Test flood_fill with an empty sheet."""
        empty_ws = self.wb.create_sheet(title="EmptySheet")
        with self.assertRaises(ValueError):
            flood_fill(empty_ws, 1, 1, empty_ws.max_row, empty_ws.max_column, set())

    def test_flood_fill_valid_area(self):
        """Test flood_fill with a valid starting point."""
        visited = set()
        min_r, max_r, min_c, max_c = flood_fill(self.ws, 1, 1, self.ws.max_row, self.ws.max_column, visited)
        self.assertEqual((min_r, max_r, min_c, max_c), (1, 2, 1, 2), "Flood fill should detect the correct table boundaries.")

class TestMetadataGeneration(unittest.TestCase):
    def setUp(self):
        """Set up test fixtures."""
        self.test_dir = "test_output"
        os.makedirs(self.test_dir, exist_ok=True)
        self.input_file = os.path.join(self.test_dir, "test.xlsx")
        self.output_files = [
            os.path.join(self.test_dir, "table1.csv"),
            os.path.join(self.test_dir, "table2.csv")
        ]
        
        # Create test files
        wb = openpyxl.Workbook()
        wb.save(self.input_file)
        for f in self.output_files:
            with open(f, 'w') as fh:
                fh.write("test,data\n1,2\n")
    
    def tearDown(self):
        """Clean up test files."""
        if os.path.exists(self.test_dir):
            for f in os.listdir(self.test_dir):
                os.unlink(os.path.join(self.test_dir, f))
            os.rmdir(self.test_dir)
    
    def test_metadata_generation(self):
        """Test DCAT metadata generation."""
        g = generate_dcat_metadata(
            self.input_file,
            self.output_files,
            base_uri="http://test.org/dataset/",
            publisher_uri="http://test.org/publisher",
            publisher_name="Test Org",
            license_uri="http://test.org/license"
        )
        
        self.assertIsInstance(g, Graph)
        turtle = g.serialize(format="turtle")
        self.assertIn("http://test.org/dataset/test", turtle)
        self.assertIn("http://test.org/publisher", turtle)
        self.assertIn("Test Org", turtle)

class TestEndToEnd(unittest.TestCase):
    def setUp(self):
        """Set up test fixtures."""
        self.test_dir = "test_output_e2e"
        os.makedirs(self.test_dir, exist_ok=True)

        # Create test Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        ws['B1'] = 'Data'
        ws['A2'] = '1'
        ws['B2'] = '2'

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        # Set both attributes to the same value
        self.excel_bytes = excel_buffer.getvalue()
        self.test_excel_bytes = self.excel_bytes

    def tearDown(self):
        """Clean up test files."""
        if os.path.exists(self.test_dir):
            for f in os.listdir(self.test_dir):
                os.unlink(os.path.join(self.test_dir, f))
            os.rmdir(self.test_dir)
    
    def test_process_excel_in_memory(self):
        """Test end-to-end processing."""
        output_files = process_excel_in_memory(self.excel_bytes, self.test_dir)
        
        self.assertTrue(output_files)
        self.assertTrue(all(os.path.exists(f) for f in output_files))
        
        # Check CSV content
        df = pd.read_csv(output_files[0])
        self.assertEqual(df.shape, (2, 2))

    def test_extract_simple_table(self):
        """Test extraction of a simple table."""
        tables = extract_tables_from_excel(self.test_excel_bytes)
        
        # Should return one table
        self.assertEqual(len(tables), 1)
        
        # Check table name and content
        name, csv_buffer = tables[0]
        self.assertEqual(name, 'Sheet')  # Default sheet name
        
        # Read CSV content
        df = pd.read_csv(csv_buffer)
        self.assertEqual(df.shape, (2, 2))  # 2x2 table
        self.assertEqual(df.iloc[0, 0], 'Test')

if __name__ == '__main__':
    unittest.main()
