"""Core functionality for Excel table detection and processing."""

from typing import List, Tuple, Set, Optional, Dict, Any
from collections import deque
from io import BytesIO
import pandas as pd
import openpyxl
import os
import re
from tqdm import tqdm

# --- Helper Functions ---
def is_cell_effectively_empty(value) -> bool:
    """Check if a cell's value is None or a string containing only whitespace."""
    if value is None:
        return True
    if isinstance(value, str) and value.isspace(): 
        return True
    return False

def is_numeric_value(value) -> bool:
    """Check if a cell's value represents a number (integer, float, or numeric string)."""
    if value is None:
        return False
    
    # If it's already a number type
    if isinstance(value, (int, float)):
        return True
    
    # If it's a string, try to parse it as a number
    if isinstance(value, str):
        value_str = value.strip()
        if not value_str:
            return False
        try:
            float(value_str)
            return True
        except ValueError:
            return False
    
    return False

def is_string_value(value) -> bool:
    """Check if a cell's value is a string type (not a number, date, time, etc.)."""
    if value is None:
        return False
    
    # Import datetime types for checking
    from datetime import datetime, date, time
    
    # If it's a date, time, or datetime, it's not a string
    if isinstance(value, (datetime, date, time)):
        return False
    
    # If it's a number, it's not a string
    if isinstance(value, (int, float)):
        return False
    
    # If it's a string, check if it's actually a number in string format
    if isinstance(value, str):
        value_str = value.strip()
        if not value_str:
            return False
        
        # Try to parse as number - if it succeeds, it's not a pure string
        try:
            float(value_str)
            return False  # It's a numeric string, not a pure string
        except ValueError:
            # It's a pure string (not numeric)
            return True
    
    # For any other type, consider it non-string
    return False

def get_merged_cell_map(sheet):
    """
    Creates a map where keys are (row, col) tuples of cells within merged ranges,
    and values are (min_row, max_row, min_col, max_col) tuples defining the extent
    of the merged range they belong to.
    """
    merged_map = {}
    if hasattr(sheet, 'merged_cells') and sheet.merged_cells:
        for merged_range in sheet.merged_cells.ranges:
            for r_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for c_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(r_idx, c_idx)] = (
                        merged_range.min_row, merged_range.max_row,
                        merged_range.min_col, merged_range.max_col
                    )
    return merged_map

def flood_fill_merged_aware(
    sheet,
    start_row: int, start_col: int,
    max_sheet_row: int, max_sheet_col: int,
    visited: Set[Tuple[int, int]],
    merged_cell_map: dict
) -> Optional[Tuple[int, int, int, int]]:
    """
    Perform flood fill to find connected non-effectively-empty cells,
    aware of merged cell ranges.
    The 'visited' set is modified by this function.
    """
    start_check_r, start_check_c = start_row, start_col
    start_mc_extent = merged_cell_map.get((start_row, start_col))
    if start_mc_extent:
        start_check_r, start_check_c = start_mc_extent[0], start_mc_extent[2]  # Top-left anchor

    if is_cell_effectively_empty(sheet.cell(start_check_r, start_check_c).value):
        return None  # Starting cell (or its anchor) is empty

    queue = deque([(start_row, start_col)])
    
    current_min_r, current_max_r = start_row, start_row
    current_min_c, current_max_c = start_col, start_col
    
    cells_in_this_region = set()

    while queue:
        r, c = queue.popleft()

        if (r, c) in cells_in_this_region or (r, c) in visited:
            continue
        
        val_check_r, val_check_c = r, c
        min_r_eff, max_r_eff, min_c_eff, max_c_eff = r, r, c, c
        
        mc_extent = merged_cell_map.get((r, c))
        if mc_extent:
            min_r_eff, max_r_eff, min_c_eff, max_c_eff = mc_extent
            val_check_r, val_check_c = mc_extent[0], mc_extent[2]  # Value is at top-left
        
        if is_cell_effectively_empty(sheet.cell(val_check_r, val_check_c).value):
            continue

        for ir in range(min_r_eff, max_r_eff + 1):
            for ic in range(min_c_eff, max_c_eff + 1):
                if (ir, ic) not in visited:
                    cells_in_this_region.add((ir, ic))
        
        current_min_r = min(current_min_r, min_r_eff)
        current_max_r = max(current_max_r, max_r_eff)
        current_min_c = min(current_min_c, min_c_eff)
        current_max_c = max(current_max_c, max_c_eff)

        if min_r_eff > 1:
            for explor_c in range(min_c_eff, max_c_eff + 1):
                nr, nc = min_r_eff - 1, explor_c
                if (nr, nc) not in visited and (nr, nc) not in cells_in_this_region:
                    n_val_r, n_val_c = nr, nc
                    n_mc_extent = merged_cell_map.get((nr, nc))
                    if n_mc_extent:
                        n_val_r, n_val_c = n_mc_extent[0], n_mc_extent[2]
                    if not is_cell_effectively_empty(sheet.cell(n_val_r, n_val_c).value):
                        queue.append((nr, nc))
        if max_r_eff < max_sheet_row:
            for explor_c in range(min_c_eff, max_c_eff + 1):
                nr, nc = max_r_eff + 1, explor_c
                if (nr, nc) not in visited and (nr, nc) not in cells_in_this_region:
                    n_val_r, n_val_c = nr, nc
                    n_mc_extent = merged_cell_map.get((nr, nc))
                    if n_mc_extent:
                        n_val_r, n_val_c = n_mc_extent[0], n_mc_extent[2]
                    if not is_cell_effectively_empty(sheet.cell(n_val_r, n_val_c).value):
                        queue.append((nr, nc))
        if min_c_eff > 1:
            for explor_r in range(min_r_eff, max_r_eff + 1):
                nr, nc = explor_r, min_c_eff - 1
                if (nr, nc) not in visited and (nr, nc) not in cells_in_this_region:
                    n_val_r, n_val_c = nr, nc
                    n_mc_extent = merged_cell_map.get((nr, nc))
                    if n_mc_extent:
                        n_val_r, n_val_c = n_mc_extent[0], n_mc_extent[2]
                    if not is_cell_effectively_empty(sheet.cell(n_val_r, n_val_c).value):
                        queue.append((nr, nc))
        if max_c_eff < max_sheet_col:
            for explor_r in range(min_r_eff, max_r_eff + 1):
                nr, nc = explor_r, max_c_eff + 1
                if (nr, nc) not in visited and (nr, nc) not in cells_in_this_region:
                    n_val_r, n_val_c = nr, nc
                    n_mc_extent = merged_cell_map.get((nr, nc))
                    if n_mc_extent:
                        n_val_r, n_val_c = n_mc_extent[0], n_mc_extent[2]
                    if not is_cell_effectively_empty(sheet.cell(n_val_r, n_val_c).value):
                        queue.append((nr, nc))
    
    if not cells_in_this_region:
        return None
    
    visited.update(cells_in_this_region)
    return current_min_r, current_max_r, current_min_c, current_max_c

def get_initial_blocks(sheet, max_row: int, max_col: int, merged_cell_map: dict) -> List[Tuple[int, int, int, int]]:
    """Detect initial table blocks, aware of merged cells."""
    visited = set()
    boxes = []
    for r_start in range(1, max_row + 1):
        for c_start in range(1, max_col + 1):
            if (r_start, c_start) in visited:
                continue

            val_check_r, val_check_c = r_start, c_start
            is_part_of_merged = False
            mc_extent = merged_cell_map.get((r_start, c_start))
            if mc_extent:
                is_part_of_merged = True
                val_check_r, val_check_c = mc_extent[0], mc_extent[2]
            
            if is_part_of_merged and (r_start != val_check_r or c_start != val_check_c):
                if (val_check_r, val_check_c) in visited and (r_start, c_start) not in visited:
                     visited.add((r_start, c_start))
                continue

            if not is_cell_effectively_empty(sheet.cell(val_check_r, val_check_c).value):
                box = flood_fill_merged_aware(sheet, r_start, c_start, max_row, max_col, visited, merged_cell_map)
                if box:
                    boxes.append(box)
            elif (r_start, c_start) not in visited:
                visited.add((r_start, c_start))

    return boxes

def merge_boxes(boxes: List[Tuple[int, int, int, int]], row_tol: int = 1, col_tol: int = 1) -> List[Tuple[int, int, int, int]]:
    """Merge nearby table boxes using tolerance parameters."""
    if not boxes:
        return []

    boxes.sort()
    merged_boxes = []

    for box in boxes:
        for i, existing in enumerate(merged_boxes):
            if (
                existing[1] + row_tol >= box[0] and existing[0] <= box[1] + row_tol and
                existing[3] + col_tol >= box[2] and existing[2] <= box[3] + col_tol
            ):
                merged_boxes[i] = (
                    min(existing[0], box[0]),
                    max(existing[1], box[1]),
                    min(existing[2], box[2]),
                    max(existing[3], box[3])
                )
                break
        else:
            merged_boxes.append(box)

    return merged_boxes

def get_used_range_effective(sheet, merged_cell_map) -> Tuple[int, int, int, int]:
    """Find the effective used range in the sheet."""
    min_r_used, max_r_used = sheet.max_row + 1, 0
    min_c_used, max_c_used = sheet.max_column + 1, 0
    found_any = False

    for r_iter in range(1, sheet.max_row + 1):
        for c_iter in range(1, sheet.max_column + 1):
            val_check_r, val_check_c = r_iter, c_iter
            mc_extent = merged_cell_map.get((r_iter, c_iter))
            if mc_extent:
                val_check_r, val_check_c = mc_extent[0], mc_extent[2]

            if not is_cell_effectively_empty(sheet.cell(val_check_r, val_check_c).value):
                found_any = True
                eff_min_r, eff_max_r, eff_min_c, eff_max_c = r_iter, r_iter, c_iter, c_iter
                if mc_extent:
                    eff_min_r, eff_max_r, eff_min_c, eff_max_c = mc_extent
                
                min_r_used = min(min_r_used, eff_min_r)
                max_r_used = max(max_r_used, eff_max_r)
                min_c_used = min(min_c_used, eff_min_c)
                max_c_used = max(max_c_used, eff_max_c)
    
    if not found_any:
        return (1, 0, 1, 0)
    return (min_r_used, max_r_used, min_c_used, max_c_used)

def attempt_horizontal_merge(sheet, blocks: List[Tuple[int, int, int, int]], 
                             max_empty_gap_cols: int = 1,
                             min_vertical_overlap_ratio: float = 0.8) -> List[Tuple[int, int, int, int]]:
    """Attempt to merge horizontally aligned blocks with significant vertical overlap."""
    if not blocks:
        return []

    current_blocks = sorted(blocks, key=lambda b: (b[0], b[2]))

    while True: 
        merged_in_this_pass = False
        next_blocks_candidate = [] 
        merged_into_another_in_pass = [False] * len(current_blocks)

        for i in range(len(current_blocks)):
            if merged_into_another_in_pass[i]: 
                continue
            
            base_block_being_built = list(current_blocks[i])

            for j in range(i + 1, len(current_blocks)):
                if merged_into_another_in_pass[j]: 
                    continue
                
                candidate_block = current_blocks[j]

                # Condition 1: Significant Vertical Overlap
                b_min_r, b_max_r = base_block_being_built[0], base_block_being_built[1]
                c_min_r, c_max_r = candidate_block[0], candidate_block[1]
                
                overlap_start_row = max(b_min_r, c_min_r)
                overlap_end_row = min(b_max_r, c_max_r)
                overlap_height = (overlap_end_row - overlap_start_row + 1) if overlap_start_row <= overlap_end_row else 0

                base_height = b_max_r - b_min_r + 1
                cand_height = c_max_r - c_min_r + 1

                is_significantly_overlapped = False
                if base_height > 0 and cand_height > 0 and overlap_height > 0:
                    if (overlap_height / base_height >= min_vertical_overlap_ratio and 
                        overlap_height / cand_height >= min_vertical_overlap_ratio):
                        is_significantly_overlapped = True
                
                is_candidate_to_right = candidate_block[2] > base_block_being_built[3]

                if is_significantly_overlapped and is_candidate_to_right:
                    gap_start_col = base_block_being_built[3] + 1
                    gap_end_col = candidate_block[2] - 1
                    num_gap_cols = (gap_end_col - gap_start_col + 1) if gap_start_col <= gap_end_col else 0

                    if 0 <= num_gap_cols <= max_empty_gap_cols:
                        gap_is_truly_empty = True
                        if num_gap_cols > 0:
                            # Check gap only within the overlapping rows
                            for r_gap in range(overlap_start_row, overlap_end_row + 1):
                                for c_gap in range(gap_start_col, gap_end_col + 1):
                                    cell_val_in_gap = sheet.cell(r_gap, c_gap).value
                                    if not is_cell_effectively_empty(cell_val_in_gap):
                                        gap_is_truly_empty = False
                                        break
                                if not gap_is_truly_empty:
                                    break
                        
                        if gap_is_truly_empty:
                            # Merge: update extents
                            base_block_being_built[0] = min(b_min_r, c_min_r) # New min_row
                            base_block_being_built[1] = max(b_max_r, c_max_r) # New max_row
                            base_block_being_built[3] = max(base_block_being_built[3], candidate_block[3]) # New max_col
                            merged_into_another_in_pass[j] = True 
                            merged_in_this_pass = True
            
            next_blocks_candidate.append(tuple(base_block_being_built))
        
        current_blocks = sorted(next_blocks_candidate, key=lambda b: (b[0], b[2]))
        
        if not merged_in_this_pass:
            break
            
    return current_blocks

def process_merged_cells(sheet) -> List[Tuple[int, int, int, int, any]]:
    """Process merged cells in the sheet."""
    return [
        (
            merged_range.min_row, merged_range.max_row,
            merged_range.min_col, merged_range.max_col,
            sheet.cell(merged_range.min_row, merged_range.min_col).value
        )
        for merged_range in sheet.merged_cells.ranges
    ]

def count_effective_content_segments(sheet, r: int, c_start: int, c_end: int, merged_cell_map: dict) -> int:
    """Counts contiguous segments of non-effectively-empty cells in a given row range, respecting merged cells."""
    segments = 0
    in_segment = False
    current_c = c_start
    while current_c <= c_end:
        val_check_r, val_check_c = r, current_c
        mc_extent = merged_cell_map.get((r, current_c))
        effective_cell_max_c = current_c

        if mc_extent:
            val_check_r, val_check_c = mc_extent[0], mc_extent[2]  # Anchor for value
            effective_cell_max_c = min(mc_extent[3], c_end)  # Merged cell ends at its own max_col or block_max_col
        
        is_empty = is_cell_effectively_empty(sheet.cell(val_check_r, val_check_c).value)

        if not is_empty and not in_segment:
            segments += 1
            in_segment = True
        elif is_empty and in_segment:
            in_segment = False        
        current_c = effective_cell_max_c + 1  # Move to the next cell after the current one (or merged segment)
    return segments


def extract_tables_from_excel(excel_bytes: bytes) -> List[Dict[str, Any]]:
    """Extract tables from Excel file, returning detailed table information."""
    excel_buffer = BytesIO(excel_bytes)
    outputs = []
    wb = openpyxl.load_workbook(filename=excel_buffer, read_only=False, data_only=True)

    # Process sheets with progress bar
    for sheet_name in tqdm(wb.sheetnames, desc="Processing Excel sheets", unit="sheet", leave=False):
        sheet = wb[sheet_name]
        current_sheet_merged_map = get_merged_cell_map(sheet)

        # 1. Initial Block Detection
        boxes = get_initial_blocks(sheet, sheet.max_row, sheet.max_column, current_sheet_merged_map)
        
        # Fallback for completely empty sheets or sheets with no detectable blocks initially
        if not boxes and sheet.max_row > 0 and sheet.max_column > 0:
            min_r, max_r, min_c, max_c = get_used_range_effective(sheet, current_sheet_merged_map)
            if max_r >= min_r and max_c >= min_c: # Check if a valid range was found
                boxes = [(min_r, max_r, min_c, max_c)]
        
        # 2. Consolidate initial fragments
        boxes = merge_boxes(boxes, row_tol=0, col_tol=0) 
        boxes.sort(key=lambda b: (b[0], b[1], b[2], b[3])) # Sort for predictable processing        # 3. Horizontal merging of adjacent blocks
        boxes = attempt_horizontal_merge(sheet, boxes, max_empty_gap_cols=2, min_vertical_overlap_ratio=0.8)
        boxes = merge_boxes(boxes, row_tol=0, col_tol=0)

        # 4. Filter trivial blocks  
        final_boxes_for_csv = []
        for r_min, r_max, c_min, c_max in boxes:
            if r_min == r_max and c_min == c_max:  # 1x1 cell block
                val_check_r, val_check_c = r_min, c_min
                mc_extent = current_sheet_merged_map.get((r_min, c_min))
                if mc_extent:
                    val_check_r, val_check_c = mc_extent[0], mc_extent[2]
                cell_value_obj = sheet.cell(val_check_r, val_check_c).value
                s_val = str(cell_value_obj if cell_value_obj is not None else "").strip()
                if not re.search(r'[a-zA-Z0-9]', s_val):
                    print(f"INFO: Skipping trivial 1x1 table at ({r_min},{c_min}) with content: '{str(cell_value_obj)}' (no alphanumeric chars)")
                    continue
            final_boxes_for_csv.append((r_min, r_max, c_min, c_max))        # 5. Process each table with new logic: merged cells → titles → headers → data        # First pass: Extract title/header info from all tables
        table_processing_results = []
        for block_idx, (min_r, max_r, min_c, max_c) in enumerate(tqdm(final_boxes_for_csv, desc=f"Processing tables in {sheet_name}", unit="table", leave=False)):
           
            # Step 1: Consolidate Merged Cells for the Current Block

            
            # Create a consolidated table where merged cell values are repeated across the entire merged range
            block_height = max_r - min_r + 1
            block_width = max_c - min_c + 1
            consolidated_block_data = []
            
            for i in range(block_height):  # for each row index within the block
                current_sheet_row = min_r + i
                current_consolidated_row = []
                
                for j in range(block_width):  # for each column index within the block
                    current_sheet_col = min_c + j
                    
                    # Check if this cell is part of a merged range
                    mc_extent = current_sheet_merged_map.get((current_sheet_row, current_sheet_col))
                    if mc_extent:
                        # This cell is part of a merged range - always get value from anchor cell
                        anchor_r, merged_max_r, anchor_c, merged_max_c = mc_extent
                        cell_value = sheet.cell(anchor_r, anchor_c).value
                        current_consolidated_row.append(cell_value)

                    else:
                        # Regular, non-merged cell
                        cell_value = sheet.cell(current_sheet_row, current_sheet_col).value
                        current_consolidated_row.append(cell_value)
                        
                
                consolidated_block_data.append(current_consolidated_row)
             
            
            if not consolidated_block_data or not any(any(str(cell).strip() for cell in row if not is_cell_effectively_empty(cell)) for row in consolidated_block_data):
              
                continue
            
            # Step 2: Extract Title Information
  
            table_title, data_after_title = extract_title_info(consolidated_block_data)
              # Step 3: Detect Header Information
      
            csv_headers, final_data_rows = detect_header_info(data_after_title)
            
            # Store the processing result
            table_processing_results.append({
                "block_info": (min_r, max_r, min_c, max_c),
                "block_idx": block_idx,
                "table_title": table_title,
                "csv_headers": csv_headers,
                "final_data_rows": final_data_rows,
                "has_data": len(final_data_rows) > 0
            })
            
       
        
        # Second pass: Handle title/header attachment for tables with no data
     
        final_table_entries = []
        
        i = 0
        while i < len(table_processing_results):
            current_result = table_processing_results[i]
            
            if current_result["has_data"]:
                # This table has data - process it normally
            
                final_table_entries.append(current_result)
                i += 1
            else:
                # This table has no data - check if we can attach it to the next table
              
                
                # Look for the next table with data
                next_data_table_idx = None
                for j in range(i + 1, len(table_processing_results)):
                    if table_processing_results[j]["has_data"]:
                        next_data_table_idx = j
                        break
                
                if next_data_table_idx is not None:
                    # Found a next table with data - attach current table's title/headers to it
                    next_result = table_processing_results[next_data_table_idx]
               
                    
                    # Combine titles
                    combined_title = None
                    if current_result["table_title"] and next_result["table_title"]:
                        combined_title = f"{current_result['table_title']} - {next_result['table_title']}"
                    elif current_result["table_title"]:
                        combined_title = current_result["table_title"]
                    elif next_result["table_title"]:
                        combined_title = next_result["table_title"]
                    
                    # Combine headers (current table's headers take precedence if both exist)
                    combined_headers = current_result["csv_headers"] if current_result["csv_headers"] else next_result["csv_headers"]
                    
                    # Create the merged result
                    merged_result = {
                        "block_info": next_result["block_info"],  # Use the data table's block info
                        "block_idx": next_result["block_idx"],
                        "table_title": combined_title,
                        "csv_headers": combined_headers,
                        "final_data_rows": next_result["final_data_rows"],
                        
                        "has_data": True
                    }
                    
              
                    final_table_entries.append(merged_result)
                    
                    # Skip both the current table (no data) and the next table (merged)
                    i = next_data_table_idx + 1
                else:
                    # No next table with data found - skip this orphaned table
 
                    i += 1
        
        # Third pass: Generate CSV files from final table entries

        table_entries = []
        for entry in final_table_entries:
            min_r, max_r, min_c, max_c = entry["block_info"]

              # Step 4: Data Processing and CSV Generation

            # Validate headers vs data columns consistency
            if entry['final_data_rows']:
                actual_columns = len(entry['final_data_rows'][0]) if entry['final_data_rows'] else 0
                detected_headers = entry['csv_headers'] if entry['csv_headers'] else []
                
                # Check if we have headers and if they match the number of columns
                if detected_headers and any(str(header).strip() for header in detected_headers):
                    # We have non-empty headers
                    if len(detected_headers) == actual_columns:                        # Headers match data columns - use them
                        df = pd.DataFrame(entry['final_data_rows'], columns=detected_headers)
                        include_header = True
                    else:
                        # Headers don't match data columns - create without headers
                        df = pd.DataFrame(entry['final_data_rows'])
                        entry['csv_headers'] = []  # Clear headers since they don't match
                        include_header = False
                else:
                    # No proper headers detected, create DataFrame without headers
                    df = pd.DataFrame(entry['final_data_rows'])
                    entry['csv_headers'] = []  # Clear headers since none were detected
                    include_header = False
            else:
                # No data rows - create empty DataFrame
                df = pd.DataFrame()
                entry['csv_headers'] = []
                include_header = False
            
            # Generate CSV
            buf = BytesIO()
            df.to_csv(buf, index=False, header=include_header)
            buf.seek(0)
            
            # Step 5: Assembling Output

            table_name_suffix = f"table_{len(table_entries) + 1}"
            table_name = f"{sheet_name}_{table_name_suffix}" 
            if len(final_table_entries) == 1:  # Only use sheet name if single table                if sheet_name not in ["Sheet", "Sheet1"] and not any(c.isdigit() for c in sheet_name):
                    table_name = sheet_name
            else:
                    table_name = f"{sheet_name}_{table_name_suffix}" if sheet_name else table_name_suffix
            if not table_name: 
                table_name = f"UnnamedSheet_table_{len(table_entries) + 1}"

            table_entries.append({
                "name": table_name, 
                "buffer": buf, 
                "headers": entry['csv_headers'], 
                "dims": entry["block_info"],                "identified_title": entry['table_title'],
                "has_original_headers": include_header
            })
            
     
        # Add all valid table entries to outputs
        outputs.extend(table_entries)

    return outputs


def extract_title_info(consolidated_data: List[List[Any]]) -> Tuple[Optional[str], List[List[Any]]]:
    """
    Extract title rows from the beginning of a table.
    Title rows are single-content rows (one non-empty, non-numeric cell) that appear before multi-content rows or data rows.
    Returns: (title_text, remaining_table_data_rows)
    """

    
    if not consolidated_data:

        return None, consolidated_data
    
    title_parts = []
    rows_consumed_by_title = 0
    
    for idx, row in enumerate(consolidated_data):

          # Count meaningful (non-empty) cells in the row
        meaningful_cells = []
        for cell_value in row:
            if not is_cell_effectively_empty(cell_value):
                cell_str = str(cell_value).strip()
                if cell_str:  # Only count if there's actual non-empty content
                    meaningful_cells.append(cell_value)  # Store original value, not string

        
        if len(meaningful_cells) == 1:
            # Single meaningful cell - check if it's a string (not numeric, date, time, etc.)
            cell_content = meaningful_cells[0]  # Use original value
            if is_string_value(cell_content):
                # Single string cell - this is a title row
                title_parts.append(str(cell_content).strip())  # Convert to string when adding
                rows_consumed_by_title += 1

            else:
                # Single non-string cell (numeric, date, time, etc.) - stop title search
       
                break
        elif len(meaningful_cells) == 0:
            # Empty row - stop title search

            break
        else:
            # Multiple meaningful cells - this could be headers or data, stop title search

            break
    
    if title_parts:
        combined_title = " - ".join(title_parts)
        remaining_rows = consolidated_data[rows_consumed_by_title:]

        return combined_title, remaining_rows
    else:

        return None, consolidated_data

def detect_header_info(data_for_headers: List[List[Any]]) -> Tuple[List[str], List[List[Any]]]:
    """
    Detect header rows using the new logic:
    - Headers are rows with multiple cells and no numeric values
    - Check subsequent rows for the same pattern until a row with a number is found or empty row
    - Concatenate multi-row headers column-wise with " - "
    Returns: (final_header_list, actual_data_rows)
    """

    if not data_for_headers:

        return [], []
    
    # Print first few rows for debugging


    
    potential_header_rows_content = []
    rows_consumed_by_header = 0
    
    for idx, row in enumerate(data_for_headers):

        
        # Count meaningful (non-empty) cells in the row
        meaningful_cells = []
        for cell_value in row:
            if not is_cell_effectively_empty(cell_value):
                cell_str = str(cell_value).strip()
                if cell_str:  # Only count if there's actual non-empty content
                    meaningful_cells.append(cell_value)
        
                    non_empty_cell_count = len(meaningful_cells)
        has_non_string = any(not is_string_value(cell) for cell in meaningful_cells)
        

        
        if non_empty_cell_count > 1 and not has_non_string:
            # Multiple string-only cells - this is a header row
            header_row = [str(cell).strip() if not is_cell_effectively_empty(cell) else "" for cell in row]
            potential_header_rows_content.append(header_row)
            rows_consumed_by_header += 1

        elif non_empty_cell_count == 0:
            # Empty row - stop header search

            break
        else:
            # Row has non-string content, or is a single cell, or other non-header condition - stop header search

            break
    
    # Process header rows
    final_header_list = []
    if potential_header_rows_content:

        
        if len(potential_header_rows_content) == 1:
            # Single header row
            final_header_list = potential_header_rows_content[0]

        else:
            # Multiple header rows - concatenate column-wise

            max_cols = max(len(r) for r in potential_header_rows_content) if potential_header_rows_content else 0
            final_header_list = [""] * max_cols
            
            for col_idx in range(max_cols):
                col_parts = []
                for h_row_idx in range(len(potential_header_rows_content)):
                    current_header_row_data = potential_header_rows_content[h_row_idx]
                    if col_idx < len(current_header_row_data):
                        cell_val_str = str(current_header_row_data[col_idx]).strip()
                        if cell_val_str:
                            col_parts.append(cell_val_str)
                
                final_header_list[col_idx] = " - ".join(col_parts) if col_parts else ""

    
    actual_data_rows = data_for_headers[rows_consumed_by_header:]

    
    return final_header_list, actual_data_rows

def process_excel_in_memory(
    excel_bytes: bytes,
    original_excel_filename: str,
    output_dir: str,
    base_uri: str = "http://example.org/dataset/",
    publisher_uri: Optional[str] = "http://example.org/publisher",
    publisher_name: Optional[str] = "Example Organization",
    license_uri: Optional[str] = "http://creativecommons.org/licenses/by/4.0/",
    existing_dataset_uri: Optional[str] = None,
    original_source_description: Optional[str] = None,
    enable_ai: bool = False,
    llm_provider: str = "openai",
    llm_api_key: Optional[str] = None,
    skip_header_ai: bool = False,
    skip_datatype_ai: bool = False
) -> Tuple[List[Dict[str, Any]], BytesIO]:
    """Process an Excel file in memory and save tables to CSV files."""
    
    extracted_tables_details = extract_tables_from_excel(excel_bytes)
    os.makedirs(output_dir, exist_ok=True)
    saved_csv_details: List[Dict[str, Any]] = []
    
    # AI processing if enabled
    if enable_ai and (not skip_header_ai or not skip_datatype_ai):
        
        try:
            from .ai_analysis import get_llm_analyzer
              # Only proceed if API key is available
            if llm_api_key:
                analyzer = get_llm_analyzer(llm_provider, llm_api_key)
            else:
                pass  # Skip AI processing
        except Exception as e:
            pass  # Continue without AI features
    
    for table_detail in extracted_tables_details:
        table_name = table_detail["name"]
        csv_buffer = table_detail["buffer"]
        output_file_path = os.path.join(output_dir, f"{table_name}.csv")
        with open(output_file_path, "wb") as f:
            f.write(csv_buffer.getvalue())
        
        saved_csv_details.append({
            "path": output_file_path, 
            "name": table_name, 
            "headers": table_detail["headers"],
            "table_title": table_detail.get("identified_title"),
            "has_original_headers": table_detail.get("has_original_headers", False)
        })    # AI processing for headers and datatypes if enabled
    if enable_ai and llm_api_key:
        try:
            from .ai_analysis import get_llm_analyzer, prepare_csv_sample_from_content, prepare_column_info_in_memory
            
            analyzer = get_llm_analyzer(llm_provider, llm_api_key)
              # Collect all tables that need AI processing
            tables_for_header_ai = []
            tables_for_datatype_ai = {}
            
            for csv_detail in tqdm(saved_csv_details, desc="Preparing AI data", unit="table", leave=False):
                csv_path = csv_detail["path"]
                table_name = csv_detail["name"]
                table_title = csv_detail.get("table_title")
                has_original_headers = csv_detail.get("has_original_headers", False)
                
                # Collect tables that need header generation
                if not skip_header_ai and not has_original_headers:
                    try:
                        with open(csv_path, 'rb') as f:
                            csv_content = f.read()
                        csv_sample = prepare_csv_sample_from_content(csv_content, max_rows=15)
                        
                        if csv_sample:                            tables_for_header_ai.append({
                                'table_name': table_name,
                                'csv_sample_data': csv_sample,
                                'table_title': table_title
                            })
                    except Exception as e:
                        pass  # Failed to prepare header AI data
                
                # Collect tables that need datatype validation
                if not skip_datatype_ai and csv_detail.get("headers"):
                    try:
                        with open(csv_path, 'rb') as f:
                            csv_content = f.read()
                        headers = csv_detail["headers"]
                        column_info = prepare_column_info_in_memory(csv_content, headers, max_sample_rows=20)
                        
                        if column_info:
                            tables_for_datatype_ai[table_name] = column_info
                    except Exception as e:
                        pass  # Failed to prepare datatype AI data
            
            # Batch process headers if any tables need it
            batch_header_results = {}
            if tables_for_header_ai:
                batch_header_results = analyzer.suggest_csv_headers_batch(tables_for_header_ai)
                  # Apply header results to CSV files
                for csv_detail in tqdm(saved_csv_details, desc="Applying AI headers", unit="table", leave=False):
                    table_name = csv_detail["name"]
                    csv_path = csv_detail["path"]
                    
                    if table_name in batch_header_results and batch_header_results[table_name]:
                        try:
                            suggested_headers = batch_header_results[table_name]
                            
                            # Read the CSV data and apply headers
                            import pandas as pd
                            df = pd.read_csv(csv_path, header=None)
                            
                            # Add headers and overwrite the CSV
                            df.columns = suggested_headers[:len(df.columns)]  # Ensure we don't have more headers than columns
                            df.to_csv(csv_path, index=False, header=True)
                              # Update the headers in saved_csv_details
                            csv_detail["headers"] = suggested_headers[:len(df.columns)]
                            csv_detail["has_original_headers"] = True
                        except Exception as e:
                            pass  # Failed to apply AI headers
              # Batch process datatypes if any tables need it
            batch_datatype_results = {}
            if tables_for_datatype_ai:
                batch_datatype_results = analyzer.suggest_column_datatypes_batch(tables_for_datatype_ai)                # Apply datatype results
                for csv_detail in tqdm(saved_csv_details, desc="Applying AI datatypes", unit="table", leave=False):
                    table_name = csv_detail["name"]
                    
                    if table_name in batch_datatype_results and batch_datatype_results[table_name]:
                        csv_detail["validated_column_types"] = batch_datatype_results[table_name]        
        except ImportError as e:
            pass  # AI modules not available
        except Exception as e:
            pass  # AI processing failed    # Generate DCAT metadata
    with tqdm(total=1, desc="Generating DCAT metadata", unit="metadata", leave=False) as pbar:
        from .metadata import generate_dcat_metadata
        metadata_graph = generate_dcat_metadata(
            original_excel_filename,
            saved_csv_details, 
            existing_dataset_uri=existing_dataset_uri,
            original_source_description=original_source_description,
            base_uri=base_uri,
            publisher_uri=publisher_uri,
            publisher_name=publisher_name,
            license_uri=license_uri,
        )
        pbar.update(1)
    
    # Serialize metadata to buffer
    with tqdm(total=1, desc="Serializing metadata", unit="buffer", leave=False) as pbar:
        metadata_buffer = BytesIO()
        metadata_graph.serialize(destination=metadata_buffer, format="turtle")
        metadata_buffer.seek(0)
        pbar.update(1)
    
    return saved_csv_details, metadata_buffer
