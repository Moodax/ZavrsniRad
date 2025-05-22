"""Core functionality for Excel table detection and processing."""

from typing import List, Tuple, Set, Optional
from collections import deque
from io import BytesIO
import pandas as pd
import openpyxl
from rdflib import Graph, Literal, URIRef, Namespace
from rdflib.namespace import DCAT, DCTERMS, FOAF, RDF, XSD
from datetime import datetime
import os
import re  # Add import for regex
from .metadata import generate_dcat_metadata

# --- New Helper Function ---
def is_cell_effectively_empty(value) -> bool:
    """Check if a cell's value is None or a string containing only whitespace."""
    if value is None:
        return True
    if isinstance(value, str) and value.isspace(): # Reverted: empty string "" is NOT considered empty
        return True
    return False

def get_merged_cell_map(sheet):
    """
    Creates a map where keys are (row, col) tuples of cells within merged ranges,
    and values are (min_row, max_row, min_col, max_col) tuples defining the extent
    of the merged range they belong to.
    """
    merged_map = {}
    if hasattr(sheet, 'merged_cells') and sheet.merged_cells:
        for merged_range in sheet.merged_cells.ranges:
            for r_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for c_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(r_idx, c_idx)] = (
                        merged_range.min_row, merged_range.max_row,
                        merged_range.min_col, merged_range.max_col
                    )
    return merged_map
# --- End New Helper Function ---

# Utility functions

def get_neighbors(r: int, c: int, max_row: int, max_col: int) -> List[Tuple[int, int]]:
    """Get valid neighboring cells."""
    return [
        (nr, nc)
        for nr, nc in [(r - 1, c), (r + 1, c), (r, c - 1), (r, c + 1)]
        if 1 <= nr <= max_row and 1 <= nc <= max_col
    ]

def flood_fill_merged_aware(
    sheet,
    start_row: int, start_col: int,
    max_sheet_row: int, max_sheet_col: int,
    visited: Set[Tuple[int, int]],
    merged_cell_map: dict
) -> Optional[Tuple[int, int, int, int]]:
    """
    Perform flood fill to find connected non-effectively-empty cells,
    aware of merged cell ranges.
    The 'visited' set is modified by this function.
    """
    start_check_r, start_check_c = start_row, start_col
    start_mc_extent = merged_cell_map.get((start_row, start_col))
    if start_mc_extent:
        start_check_r, start_check_c = start_mc_extent[0], start_mc_extent[2]  # Top-left anchor

    if is_cell_effectively_empty(sheet.cell(start_check_r, start_check_c).value):
        return None  # Starting cell (or its anchor) is empty

    queue = deque([(start_row, start_col)])
    
    current_min_r, current_max_r = start_row, start_row
    current_min_c, current_max_c = start_col, start_col
    
    cells_in_this_region = set()

    while queue:
        r, c = queue.popleft()

        if (r, c) in cells_in_this_region or (r, c) in visited:
            continue
        
        val_check_r, val_check_c = r, c
        min_r_eff, max_r_eff, min_c_eff, max_c_eff = r, r, c, c
        
        mc_extent = merged_cell_map.get((r, c))
        if mc_extent:
            min_r_eff, max_r_eff, min_c_eff, max_c_eff = mc_extent
            val_check_r, val_check_c = mc_extent[0], mc_extent[2]  # Value is at top-left
        
        if is_cell_effectively_empty(sheet.cell(val_check_r, val_check_c).value):
            continue

        for ir in range(min_r_eff, max_r_eff + 1):
            for ic in range(min_c_eff, max_c_eff + 1):
                if (ir, ic) not in visited:
                    cells_in_this_region.add((ir, ic))
        
        current_min_r = min(current_min_r, min_r_eff)
        current_max_r = max(current_max_r, max_r_eff)
        current_min_c = min(current_min_c, min_c_eff)
        current_max_c = max(current_max_c, max_c_eff)

        if min_r_eff > 1:
            for explor_c in range(min_c_eff, max_c_eff + 1):
                nr, nc = min_r_eff - 1, explor_c
                if (nr, nc) not in visited and (nr, nc) not in cells_in_this_region:
                    n_val_r, n_val_c = nr, nc
                    n_mc_extent = merged_cell_map.get((nr, nc))
                    if n_mc_extent:
                        n_val_r, n_val_c = n_mc_extent[0], n_mc_extent[2]
                    if not is_cell_effectively_empty(sheet.cell(n_val_r, n_val_c).value):
                        queue.append((nr, nc))
        if max_r_eff < max_sheet_row:
            for explor_c in range(min_c_eff, max_c_eff + 1):
                nr, nc = max_r_eff + 1, explor_c
                if (nr, nc) not in visited and (nr, nc) not in cells_in_this_region:
                    n_val_r, n_val_c = nr, nc
                    n_mc_extent = merged_cell_map.get((nr, nc))
                    if n_mc_extent:
                        n_val_r, n_val_c = n_mc_extent[0], n_mc_extent[2]
                    if not is_cell_effectively_empty(sheet.cell(n_val_r, n_val_c).value):
                        queue.append((nr, nc))
        if min_c_eff > 1:
            for explor_r in range(min_r_eff, max_r_eff + 1):
                nr, nc = explor_r, min_c_eff - 1
                if (nr, nc) not in visited and (nr, nc) not in cells_in_this_region:
                    n_val_r, n_val_c = nr, nc
                    n_mc_extent = merged_cell_map.get((nr, nc))
                    if n_mc_extent:
                        n_val_r, n_val_c = n_mc_extent[0], n_mc_extent[2]
                    if not is_cell_effectively_empty(sheet.cell(n_val_r, n_val_c).value):
                        queue.append((nr, nc))
        if max_c_eff < max_sheet_col:
            for explor_r in range(min_r_eff, max_r_eff + 1):
                nr, nc = explor_r, max_c_eff + 1
                if (nr, nc) not in visited and (nr, nc) not in cells_in_this_region:
                    n_val_r, n_val_c = nr, nc
                    n_mc_extent = merged_cell_map.get((nr, nc))
                    if n_mc_extent:
                        n_val_r, n_val_c = n_mc_extent[0], n_mc_extent[2]
                    if not is_cell_effectively_empty(sheet.cell(n_val_r, n_val_c).value):
                        queue.append((nr, nc))
    
    if not cells_in_this_region:
        return None
    
    visited.update(cells_in_this_region)
    return current_min_r, current_max_r, current_min_c, current_max_c

def get_initial_blocks(sheet, max_row: int, max_col: int, merged_cell_map: dict) -> List[Tuple[int, int, int, int]]:
    """Detect initial table blocks, aware of merged cells."""
    visited = set()
    boxes = []
    for r_start in range(1, max_row + 1):
        for c_start in range(1, max_col + 1):
            if (r_start, c_start) in visited:
                continue

            val_check_r, val_check_c = r_start, c_start
            is_part_of_merged = False
            mc_extent = merged_cell_map.get((r_start, c_start))
            if mc_extent:
                is_part_of_merged = True
                val_check_r, val_check_c = mc_extent[0], mc_extent[2]
            
            if is_part_of_merged and (r_start != val_check_r or c_start != val_check_c):
                if (val_check_r, val_check_c) in visited and (r_start, c_start) not in visited:
                     visited.add((r_start, c_start))
                continue

            if not is_cell_effectively_empty(sheet.cell(val_check_r, val_check_c).value):
                box = flood_fill_merged_aware(sheet, r_start, c_start, max_row, max_col, visited, merged_cell_map)
                if box:
                    boxes.append(box)
            elif (r_start, c_start) not in visited:
                visited.add((r_start, c_start))

    return boxes

def merge_boxes(boxes: List[Tuple[int, int, int, int]], row_tol: int = 1, col_tol: int = 1) -> List[Tuple[int, int, int, int]]:
    """Merge nearby table boxes using tolerance parameters."""
    if not boxes:
        return []

    boxes.sort()
    merged_boxes = []

    for box in boxes:
        for i, existing in enumerate(merged_boxes):
            if (
                existing[1] + row_tol >= box[0] and existing[0] <= box[1] + row_tol and
                existing[3] + col_tol >= box[2] and existing[2] <= box[3] + col_tol
            ):
                merged_boxes[i] = (
                    min(existing[0], box[0]),
                    max(existing[1], box[1]),
                    min(existing[2], box[2]),
                    max(existing[3], box[3])
                )
                break
        else:
            merged_boxes.append(box)

    return merged_boxes

def get_used_range_effective(sheet, merged_cell_map) -> Tuple[int, int, int, int]:
    """Find the effective used range in the sheet."""
    min_r_used, max_r_used = sheet.max_row + 1, 0
    min_c_used, max_c_used = sheet.max_column + 1, 0
    found_any = False

    for r_iter in range(1, sheet.max_row + 1):
        for c_iter in range(1, sheet.max_column + 1):
            val_check_r, val_check_c = r_iter, c_iter
            mc_extent = merged_cell_map.get((r_iter, c_iter))
            if mc_extent:
                val_check_r, val_check_c = mc_extent[0], mc_extent[2]

            if not is_cell_effectively_empty(sheet.cell(val_check_r, val_check_c).value):
                found_any = True
                eff_min_r, eff_max_r, eff_min_c, eff_max_c = r_iter, r_iter, c_iter, c_iter
                if mc_extent:
                    eff_min_r, eff_max_r, eff_min_c, eff_max_c = mc_extent
                
                min_r_used = min(min_r_used, eff_min_r)
                max_r_used = max(max_r_used, eff_max_r)
                min_c_used = min(min_c_used, eff_min_c)
                max_c_used = max(max_c_used, eff_max_c)
    
    if not found_any:
        return (1, 0, 1, 0)
    return (min_r_used, max_r_used, min_c_used, max_c_used)

def attempt_horizontal_merge(sheet, blocks: List[Tuple[int, int, int, int]], 
                             max_empty_gap_cols: int = 1,
                             min_vertical_overlap_ratio: float = 0.8) -> List[Tuple[int, int, int, int]]:
    """Attempt to merge horizontally aligned blocks with significant vertical overlap."""
    if not blocks:
        return []

    current_blocks = sorted(blocks, key=lambda b: (b[0], b[2]))

    while True: 
        merged_in_this_pass = False
        next_blocks_candidate = [] 
        indices_to_process_in_pass = list(range(len(current_blocks)))
        merged_into_another_in_pass = [False] * len(current_blocks)

        for i in indices_to_process_in_pass:
            if merged_into_another_in_pass[i]: 
                continue
            
            base_block_being_built = list(current_blocks[i])

            for j in range(i + 1, len(current_blocks)):
                if merged_into_another_in_pass[j]: 
                    continue
                
                candidate_block = current_blocks[j]

                # Condition 1: Significant Vertical Overlap
                b_min_r, b_max_r = base_block_being_built[0], base_block_being_built[1]
                c_min_r, c_max_r = candidate_block[0], candidate_block[1]
                
                overlap_start_row = max(b_min_r, c_min_r)
                overlap_end_row = min(b_max_r, c_max_r)
                overlap_height = (overlap_end_row - overlap_start_row + 1) if overlap_start_row <= overlap_end_row else 0

                base_height = b_max_r - b_min_r + 1
                cand_height = c_max_r - c_min_r + 1

                is_significantly_overlapped = False
                if base_height > 0 and cand_height > 0 and overlap_height > 0:
                    if (overlap_height / base_height >= min_vertical_overlap_ratio and 
                        overlap_height / cand_height >= min_vertical_overlap_ratio):
                        is_significantly_overlapped = True
                
                is_candidate_to_right = candidate_block[2] > base_block_being_built[3]

                if is_significantly_overlapped and is_candidate_to_right:
                    gap_start_col = base_block_being_built[3] + 1
                    gap_end_col = candidate_block[2] - 1
                    num_gap_cols = (gap_end_col - gap_start_col + 1) if gap_start_col <= gap_end_col else 0

                    if 0 <= num_gap_cols <= max_empty_gap_cols:
                        gap_is_truly_empty = True
                        if num_gap_cols > 0:
                            # Check gap only within the overlapping rows
                            for r_gap in range(overlap_start_row, overlap_end_row + 1):
                                for c_gap in range(gap_start_col, gap_end_col + 1):
                                    cell_val_in_gap = sheet.cell(r_gap, c_gap).value
                                    if not is_cell_effectively_empty(cell_val_in_gap):
                                        gap_is_truly_empty = False
                                        break
                                if not gap_is_truly_empty:
                                    break
                        
                        if gap_is_truly_empty:
                            # Merge: update extents
                            base_block_being_built[0] = min(b_min_r, c_min_r) # New min_row
                            base_block_being_built[1] = max(b_max_r, c_max_r) # New max_row
                            base_block_being_built[3] = max(base_block_being_built[3], candidate_block[3]) # New max_col
                            merged_into_another_in_pass[j] = True 
                            merged_in_this_pass = True
            
            next_blocks_candidate.append(tuple(base_block_being_built))
        
        current_blocks = sorted(next_blocks_candidate, key=lambda b: (b[0], b[2]))
        
        if not merged_in_this_pass:
            break
            
    return current_blocks

def process_merged_cells(sheet) -> List[Tuple[int, int, int, int, any]]:
    """Process merged cells in the sheet."""
    return [
        (
            merged_range.min_row, merged_range.max_row,
            merged_range.min_col, merged_range.max_col,
            sheet.cell(merged_range.min_row, merged_range.min_col).value
        )
        for merged_range in sheet.merged_cells.ranges
    ]

def count_effective_content_segments(sheet, r: int, c_start: int, c_end: int, merged_cell_map: dict) -> int:
    """Counts contiguous segments of non-effectively-empty cells in a given row range, respecting merged cells."""
    segments = 0
    in_segment = False
    current_c = c_start
    while current_c <= c_end:
        val_check_r, val_check_c = r, current_c
        mc_extent = merged_cell_map.get((r, current_c))
        effective_cell_max_c = current_c

        if mc_extent:
            val_check_r, val_check_c = mc_extent[0], mc_extent[2]  # Anchor for value
            effective_cell_max_c = min(mc_extent[3], c_end)  # Merged cell ends at its own max_col or block_max_col
        
        is_empty = is_cell_effectively_empty(sheet.cell(val_check_r, val_check_c).value)

        if not is_empty and not in_segment:
            segments += 1
            in_segment = True
        elif is_empty and in_segment:
            in_segment = False
        
        current_c = effective_cell_max_c + 1  # Move to the next cell after the current one (or merged segment)
    return segments

def extract_tables_from_excel(excel_bytes: bytes) -> List[Tuple[str, BytesIO]]:
    excel_buffer = BytesIO(excel_bytes)
    outputs = []
    wb = openpyxl.load_workbook(filename=excel_buffer, read_only=False, data_only=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        current_sheet_merged_map = get_merged_cell_map(sheet)

        # 1. Initial Block Detection
        boxes = get_initial_blocks(sheet, sheet.max_row, sheet.max_column, current_sheet_merged_map)
        
        # Fallback for completely empty sheets or sheets with no detectable blocks initially
        if not boxes and sheet.max_row > 0 and sheet.max_column > 0:
            min_r, max_r, min_c, max_c = get_used_range_effective(sheet, current_sheet_merged_map)
            if max_r >= min_r and max_c >= min_c: # Check if a valid range was found
                boxes = [(min_r, max_r, min_c, max_c)]
        
        # 2. Consolidate initial fragments
        boxes = merge_boxes(boxes, row_tol=0, col_tol=0) 
        boxes.sort(key=lambda b: (b[0], b[1], b[2], b[3])) # Sort for predictable processing

        # 3. Existing Title Identification (for already separated 1-row blocks above data)
        title_blocks = []
        data_contender_blocks = [] # Blocks that are not titles themselves yet
        processed_indices = [False] * len(boxes)

        for i in range(len(boxes)):
            if processed_indices[i]:
                continue

            box_i = boxes[i]
            is_potential_title = (box_i[1] - box_i[0] + 1) == 1 # 1 row high
            
            if is_potential_title:
                # Check for alphanumeric content in the potential title
                title_anchor_r, title_anchor_c = box_i[0], box_i[2]
                mc_title_extent = current_sheet_merged_map.get((title_anchor_r, title_anchor_c))
                if mc_title_extent:
                    title_anchor_r, title_anchor_c = mc_title_extent[0], mc_title_extent[2]
                title_val_obj = sheet.cell(title_anchor_r, title_anchor_c).value
                title_s_val = str(title_val_obj if title_val_obj is not None else "").strip()
                if not re.search(r'[a-zA-Z0-9]', title_s_val):
                    is_potential_title = False # Not a title if no alphanumeric chars

            found_data_below = False
            if is_potential_title:
                for j in range(len(boxes)):
                    if i == j or processed_indices[j]: 
                        continue
                    box_j = boxes[j]
                    if box_j[0] == box_i[1] + 1 and (box_j[1] - box_j[0] + 1) > 1:
                        overlap_min_c = max(box_i[2], box_j[2])
                        overlap_max_c = min(box_i[3], box_j[3])
                        if overlap_min_c <= overlap_max_c: 
                            title_blocks.append(box_i)
                            processed_indices[i] = True
                            # box_j will be added to data_contender_blocks if not processed itself
                            found_data_below = True
                            break 
            
            if not processed_indices[i]: 
                data_contender_blocks.append(box_i)
                processed_indices[i] = True
        
        # Add any blocks that were identified as data for titles, or were not titles, to data_contender_blocks
        # This logic needs to be careful not to add box_j if it was already processed.
        # A simpler way: all non-title_blocks are data_contenders initially.
        temp_data_blocks = []
        is_title_flags = [False] * len(boxes)
        for i, box_i in enumerate(boxes):
            is_this_box_a_title = False
            if (box_i[1] - box_i[0] + 1) == 1: # 1 row high
                title_anchor_r, title_anchor_c = box_i[0], box_i[2]
                mc_title_extent = current_sheet_merged_map.get((title_anchor_r, title_anchor_c))
                if mc_title_extent: title_anchor_r, title_anchor_c = mc_title_extent[0], mc_title_extent[2]
                title_val_obj = sheet.cell(title_anchor_r, title_anchor_c).value
                title_s_val = str(title_val_obj if title_val_obj is not None else "").strip()
                if re.search(r'[a-zA-Z0-9]', title_s_val):
                    for j, box_j in enumerate(boxes):
                        if i == j: continue
                        if box_j[0] == box_i[1] + 1 and (box_j[1] - box_j[0] + 1) > 1:
                            overlap_min_c = max(box_i[2], box_j[2]); overlap_max_c = min(box_i[3], box_j[3])
                            if overlap_min_c <= overlap_max_c:
                                title_blocks.append(box_i); is_this_box_a_title = True; break
            if not is_this_box_a_title:
                temp_data_blocks.append(box_i)
        data_blocks = temp_data_blocks

        # 4. Process Data Blocks for horizontal merging
        data_blocks = attempt_horizontal_merge(sheet, data_blocks, max_empty_gap_cols=2, min_vertical_overlap_ratio=0.8)
        data_blocks = merge_boxes(data_blocks, row_tol=0, col_tol=0)

        # 5. Combine and Filter Trivial
        all_identified_blocks = title_blocks + data_blocks
        all_identified_blocks.sort(key=lambda b: (b[0], b[1], b[2], b[3]))

        final_boxes_for_csv = []
        for r_min, r_max, c_min, c_max in all_identified_blocks:
            if r_min == r_max and c_min == c_max:  # 1x1 cell block
                val_check_r, val_check_c = r_min, c_min
                mc_extent = current_sheet_merged_map.get((r_min, c_min))
                if mc_extent:
                    val_check_r, val_check_c = mc_extent[0], mc_extent[2]
                cell_value_obj = sheet.cell(val_check_r, val_check_c).value
                s_val = str(cell_value_obj if cell_value_obj is not None else "").strip()
                if not re.search(r'[a-zA-Z0-9]', s_val):
                    print(f"INFO: Skipping trivial 1x1 table at ({r_min},{c_min}) with content: '{str(cell_value_obj)}' (no alphanumeric chars)")
                    continue
            final_boxes_for_csv.append((r_min, r_max, c_min, c_max))
        
        # 6. Generate CSVs
        sheet_merged_cell_data_cache = process_merged_cells(sheet)
        for idx, (min_r, max_r, min_c, max_c) in enumerate(final_boxes_for_csv, start=1):
            data = [
                [
                    next((m_value for m_min_r_mc, m_max_r_mc, m_min_c_mc, m_max_c_mc, m_value in sheet_merged_cell_data_cache
                          if m_min_r_mc <= row <= m_max_r_mc and m_min_c_mc <= col <= m_max_c_mc),
                         sheet.cell(row, col).value or "")
                    for col in range(min_c, max_c + 1)
                ]
                for row in range(min_r, max_r + 1)
            ]

            if any(any(cell_val for cell_val in data_row) for data_row in data):
                df = pd.DataFrame(data)
                buf = BytesIO()
                df.to_csv(buf, index=False, header=False)
                buf.seek(0)

                table_name = f"{sheet_name}_table_{idx}" if len(final_boxes_for_csv) > 1 else sheet_name
                outputs.append((table_name, buf))

    return outputs

def process_excel_in_memory(
    excel_bytes: bytes,
    original_excel_filename: str,
    output_dir: str,
    base_uri: str = "http://example.org/dataset/",
    publisher_uri: str = "http://example.org/publisher",
    publisher_name: str = "Example Organization",
    license_uri: str = "http://creativecommons.org/licenses/by/4.0/",
) -> Tuple[List[str], BytesIO]:
    """Process an Excel file in memory and save tables to CSV files.

    Args:
        excel_bytes: The raw bytes of the Excel file.
        original_excel_filename: The original filename of the Excel file.
        output_dir: The directory where CSV files will be saved.
        base_uri: Base URI for the dataset.
        publisher_uri: URI of the publisher.
        publisher_name: Name of the publisher.
        license_uri: URI of the license.

    Returns:
        A tuple containing:
        - List of paths to the saved CSV files
        - BytesIO containing the DCAT metadata
    """
    outputs = []

    tables = extract_tables_from_excel(excel_bytes)

    os.makedirs(output_dir, exist_ok=True)

    for table_name, csv_buffer in tables:
        output_file = os.path.join(output_dir, f"{table_name}.csv")
        with open(output_file, "wb") as f:
            f.write(csv_buffer.getvalue())
        outputs.append(output_file)

    metadata_graph = generate_dcat_metadata(
        original_excel_filename,
        [os.path.basename(f) for f in outputs],
        base_uri=base_uri,
        publisher_uri=publisher_uri,
        publisher_name=publisher_name,
        license_uri=license_uri,
    )

    metadata_buffer = BytesIO()
    metadata_graph.serialize(destination=metadata_buffer, format="turtle")
    metadata_buffer.seek(0)

    return outputs, metadata_buffer